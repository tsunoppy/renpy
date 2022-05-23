#! /Users/tsuno/.pyenv/shims/python3
# -*- coding: utf-8 -*-

import math
import openpyxl
from openpyxl.utils import get_column_letter # 列幅の指定 2020/05/27
import pandas as pd
import numpy as np
import col
import os
import shutil
import report

########################################################################
# Main program
########################################################################
class MnGen():

    ########################################################################
    # init data
    def __init__(self,input_path):


        print("--------------------")
        print("Start Main")
        print("--------------------")

        """
        if os.path.exists('./db'):
            shutil.rmtree('./db')
            os.mkdir('./db')
            print('remove ./db ^v^!')
        else:
            os.mkdir('./db')
            print('none *db')
        """

        ####################
        # path info.
        self.inp_path = input_path
        self.home_dir = os.path.dirname(self.inp_path)

    ########################################################################
    # read cntl data
    def read_cntl(self):

        df_cntl = pd.read_excel(self.inp_path, sheet_name="CNTL",\
                                engine='openpyxl')
        print("--------------------")
        print("CNTL")
        print("--------------------")
        print(df_cntl)

        # grobal var as following
        self.out_path    = self.home_dir + "/" + df_cntl.iloc[0,0]
        self.view_path   = self.home_dir + "/" + df_cntl.iloc[1,0]
        self.report_path = self.home_dir + "/" + df_cntl.iloc[2,0]
        self.report_title = str(df_cntl.iloc[2,1])
        #
        self.ndiv     = int(df_cntl.iloc[0,1])
        self.mdmax    = df_cntl.iloc[0,2]
        self.ndmin    = df_cntl.iloc[0,3]
        self.ndmax    = df_cntl.iloc[0,4]

        ####################
        # clear outputdata

        if os.path.exists(self.out_path):
            os.remove(self.out_path)
            #print('remove',self.out_path,'^v^!')
            print('exists,',self.out_path,'^v^!')

        else:
            print('none',self.out_path)
        wb = openpyxl.Workbook()
        wb.save(self.out_path)

    ########################################################################
    # read_column data
    def read_column(self):

        # read data
        ####################
        df_col = pd.read_excel(self.inp_path, sheet_name="COLUMN",\
                               engine='openpyxl' )
        # skip if comment symbol "*" is in st.
        df_col = df_col[ df_col['st'] != "*"]
        print("--------------------")
        print("COLUMN DATA")
        print("--------------------")
        print(df_col.head())

        # data making
        self.st   = df_col.iloc[:,0].astype(int).values
        self.symb = df_col.iloc[:,1].values
        self.fc   = df_col.iloc[:,2].values
        self.fy   = df_col.iloc[:,3].values
        self.b    = df_col.iloc[:,4].values
        self.d    = df_col.iloc[:,5].values
        self.dia  = df_col.iloc[:,6].astype(int).values
        #
        tmpnx = df_col[['nx1','nx2','nx3','nx4','nx5']]
        tmpny = df_col[['ny1','ny2','ny3','ny4','ny5']]
        tmpdtx = df_col[['dtx1','dtx2','dtx3','dtx4','dtx5']]
        tmpdty = df_col[['dty1','dty2','dty3','dty4','dty5']]
        #
        #print(df_col[['nx1','nx2','nx3','nx4','nx5']])
        #print( nx[ nx != '*' ])
        #print(df_col.head())
        #print(df_col)
        #print(tmpnx.head())
        #
        self.nx = []
        self.ny = []
        self.dtx = []
        self.dty = []
        for i in range(0,len(tmpnx)):
            self.nx.append( tmpnx.iloc[i,:].dropna().astype(int).values )
            self.ny.append( tmpny.iloc[i,:].dropna().astype(int).values)
            self.dtx.append( tmpdtx.iloc[i,:].dropna().values)
            self.dty.append( tmpdty.iloc[i,:].dropna().values)

        self.name = []
        for i in range(0,len(self.fc)):
            self.name.append( str(self.st[i]) + str(self.symb[i]) )

    ########################################################################
    # make datasets for deep learning
    def solve_deep(self):

        print("---------------------------------")
        print(" make datasets for deep learning ")
        print("---------------------------------")

        # load exel file
        ####################
        wb = openpyxl.load_workbook(self.out_path)

        # make mn data
        ####################
        self.obj = [] # column object

        # parameter for the deep learning
        fc_d  = []
        fy_d  = []
        b_d   = []
        d_d   = []
        szx_d = []
        szy_d = []
        ag_d  = []
        nu_d  = []
        mu_d  = []

        for i in range(0,len(self.fc)):
            self.obj.append(\
                            col.col(\
                               self.fc[i],self.fy[i],self.b[i],self.d[i],\
                                    self.nx[i],self.ny[i],\
                                    self.dtx[i],self.dty[i],self.dia[i])\
                            )
            #self.obj[i].mn_result_xlsx(self.ndiv,self.out_path,self.name[i],wb)
            six,siy,szx,szy,ag =\
                self.obj[i].feature()
            nu, mu = self.obj[i].mnuGen("X",self.ndiv)
            nu = nu*10**3/(self.b[i] * self.d[i])
            mu = mu*10**6/(self.b[i]**2 * self.d[i] )

            #print(six,siy,szx,szy,ag)
            #for i in range(0,len(nu)):
            #    print(nu[i],mu[i])

            for j in range(0,len(nu)):
                fc_d.append( self.fc[i] )
                fy_d.append( self.fy[i] )
                b_d.append(  self.b[i]  )
                d_d.append(  self.d[i]  )
                szx_d.append( szx*6/(self.b[i]**2*self.d[i])  )
                szy_d.append( szy*6/(self.b[i]*self.d[i]**2)  )
                ag_d.append(  ag/(self.b[i]*self.d[i]) )
            #nu_d.append(nu * 10**3 /(self.b[i] * self.d[i]))
            #mu_d.append(mu * 10**6 /(self.b[i]**2 * self.d[i]))
            nu_d = np.hstack( (np.array(nu_d), nu ))
            mu_d = np.hstack( (np.array(mu_d), mu ))

            #print(nu_d)

        df_deep = pd.DataFrame({'fc':fc_d,
                                'fy':fy_d,
                                'b':b_d,
                                'd':d_d,
                                'szx':szx_d,
                                'szy':szy_d,
                                'ag':ag_d,
                                'nu':nu_d,
                                'mu':mu_d})
        print(df_deep)

        df_deep.to_csv('dl.dat')
            #print('mux=',self.obj[i].mnuaci_twoM("X",0.0))
            #print('muy=',self.obj[i].mnuaci_twoM("Y",0.0))

            #print("mn generate ---- No.",i,self.name[i])
        import matplotlib.pyplot as plt
        fig = plt.figure(figsize=(6,6),tight_layout=True)
        ax = fig.add_subplot()
        ax.scatter(df_deep['mu'],df_deep['nu'],s=5)
        plt.show()


        wb.save(self.out_path)

    ########################################################################
    # solve
    def solve(self):

        print("--------------------")
        print(" generate MN-data")
        print("--------------------")

        # load exel file
        ####################
        wb = openpyxl.load_workbook(self.out_path)

        # make mn data
        ####################
        self.obj = [] # column object
        for i in range(0,len(self.fc)):
            self.obj.append(\
                            col.col(\
                               self.fc[i],self.fy[i],self.b[i],self.d[i],\
                                    self.nx[i],self.ny[i],\
                                    self.dtx[i],self.dty[i],self.dia[i])\
                            )
            self.obj[i].mn_result_xlsx(self.ndiv,self.out_path,self.name[i],wb)

            #print('mux=',self.obj[i].mnuaci_twoM("X",0.0))
            #print('muy=',self.obj[i].mnuaci_twoM("Y",0.0))

            print("mn generate ---- No.",i,self.name[i])

        wb.save(self.out_path)

        # make mn fig
        ####################
        for i in range(0,len(self.fc)):
            print("draw section ------",self.name[i])
            self.obj[i].make_model_fig(self.name[i])

    ########################################################################
    # make report
    def make_report(self):
        # cntl: control file
        # no use!!!!!!
        # make report
        cntl = "./"
        num = len(self.fc)
        pathname = self.report_path
        title = "mn"
        obj = report.Report(cntl,self.name)
        obj.create_pdf(num,pathname,title)

    ########################################################################
    # read calc
    def read_calc(self):

        ########################################################################
        # read_load
        df_cal = pd.read_excel(self.inp_path, sheet_name="CALC",\
                               engine='openpyxl')

        print(df_cal.head())

        calc_st  =  df_cal.iloc[:,0].values
        calc_sym =  df_cal.iloc[:,1].values
        calc_xnum = df_cal.iloc[:,2].values
        nend = 4+3*int(calc_xnum[0])
        print( df_cal.iloc[:,4:nend] )
        print( 'N=',df_cal['N_x1'].values )
        print( 'M=',df_cal['M_x1'].values )

        print("--------------------")
        print("CALC")
        print("--------------------")
        #print(calc_sym)

        name_cal = []
        for i in range(0,len(calc_sym)):
            if calc_st[i] != "*":
                name_cal.append( str(calc_st[i])+str(calc_sym[i]) )
            else:
                name_cal.append( "*" )
        #print( name_cal )

        """
        for i in range(0,len(name_cal)):
            if name_cal[i] != "*":
                ind = self.name.index(name_cal[i])
                # mn curve fig
                #self.obj[ind].mn_result_xlsx(self.ndiv,self.out_path,self.name[ind])
                # mn curve fig
                self.obj[ind].make_fig(self.out_path,self.name[ind])
        """

        # test
        coldata = []
        for i in range(0,len(name_cal)):
            print("mn drawing generate ---- No.",i,name_cal[i])
            if name_cal[i] != "*":
                obj = col.Aft_mn(self.out_path,name_cal[i],"XY",\
                           self.mdmax,self.ndmin,self.ndmax)

                # get stress by excel
                ## x-dir.
                xnum = int(df_cal.loc[i,'xnum'])
                xtitle = []
                nux = []
                mux = []
                if xnum != 0:
                    xtitle.append( str(df_cal.loc[i,'load_x1']) )
                    nux.append( float( df_cal.loc[i,'N_x1'] ) )
                    mux.append( float( df_cal.loc[i,'M_x1'] ) )
                    if xnum == 2:
                        xtitle.append( str(df_cal.loc[i,'load_x2']) )
                        nux.append( float( df_cal.loc[i,'N_x2'] ) )
                        mux.append( float( df_cal.loc[i,'M_x2'] ) )
                print( 'xnum, Nux, Mux=', xnum, nux, mux )

                ## y-dir.
                ynum = int(df_cal.loc[i,'ynum'])
                ytitle = []
                nuy = []
                muy = []
                if ynum != 0:
                    ytitle.append( str(df_cal.loc[i,'load_y1']) )
                    nuy.append( float( df_cal.loc[i,'N_y1'] ) )
                    muy.append( float( df_cal.loc[i,'M_y1'] ) )
                    if xnum == 2:
                        ytitle.append( str(df_cal.loc[i,'load_y2']) )
                        nuy.append( float( df_cal.loc[i,'N_y2'] ) )
                        muy.append( float( df_cal.loc[i,'M_y2'] ) )
                print( 'ynum, Nuy, Muy=', ynum, nuy, muy )

                # make mn figure by aftFib
                obj.make_fig(xnum,xtitle,nux,\
                             mux,ynum,ytitle,nuy,muy)
                coldata.append(obj.spec())
            else:
                coldata.append("*")

        num = len(name_cal)
        cntl = "./"
        pathname = self.report_path
        title = self.report_title

        rep = report.Report(cntl,name_cal,coldata,calc_st)
        rep.create_pdf(num,pathname,title)

########################################################################
# end of class

if __name__ == '__main__':
    input_path = "./input_mngen/data.xlsx"
    #input_path = "./input_mngen/before.xlsx"
    #input_path = "./input_mngen/after.xlsx"
    #input_path = "./test/data.xlsx"
    obj = MnGen(input_path)
    obj.read_cntl()
    # they are sets
    obj.read_column()
    #
    obj.solve()
    obj.read_calc()
    #or for deep learning
    #obj.solve_deep()
    ######obj.make_report() # no use

