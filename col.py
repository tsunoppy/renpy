#! /Users/tsuno/.pyenv/shims/python3
# -*- coding: utf-8 -*-

import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.ticker import ScalarFormatter

import openpyxl
#from openpyxl.utils import get_column_letter # 列幅の指定 2020/05/27
import numpy as np
import pandas as pd
import rc

########################################################################
# 柱、MN計算
class col():

    def __init__(self,fc,fy,b,dd,nx,ny,dtx,dty,dia,\
                 lfc,sfc,lfy,sfy,alpha):
        # 2022.05.31 addtional for permissible capacity
        # lfc,sf,lfy,sfy/ long & short term permissible stress
        # alpha: multiple factor to the fy
        self.lfc = lfc
        self.sfc = sfc
        self.lfy = lfy
        self.sfy = sfy
        self.alpha = alpha

        # set parameter
        self.eps = 10 ** (-2)

        self.fc  = fc
        self.fy  = fy * self.alpha
        self.b   = b
        self.dd  = dd

        self.nx   = np.array(nx)
        self.ny   = np.array(ny)
        self.dtx  = np.array(dtx)
        self.dty  = np.array(dty)
        self.dia  = dia

        # initial cal
        self.area = rc.rfc().rA(dia) # bar area

        # make bar postion
        self.xs = []
        self.ys = []
        for i in range(0,len(nx)):
            self.bar_pos(nx[i],ny[i],dtx[i],dty[i])

        # calculation of steel bar inersia
        xb = np.array(self.xs) - self.b
        yb = np.array(self.ys) - self.dd
        six = np.sum(xb**2) * self.area
        siy = np.sum(yb**2) * self.area
        szx = six / ( np.max(xb) - np.min(xb) )
        szy = siy / ( np.max(yb) - np.min(yb) )
        #print(six,siy)
        #print(szx,szy)
        self.six = six
        self.siy = siy
        self.szx = szx
        self.szy = szy

        """
        print(self.xs,self.ys)
        plt.scatter(self.xs,self.ys)
        plt.show()
        """
        self.ag = len(self.xs) * self.area
        #print(self.atx,self.nwx,self.aty,self.nwy)

        # yang modulus
        if self.fc <= 27: self.yn = 15.0
        elif 27 < self.fc and self.fc <= 36: self.yn = 13.0
        elif 48 < self.fc and self.fc <= 48: self.yn = 11.0
        else: self.yn = 9.0

        # aci parameter
        self.eu = 0.003
        if fc > 28.0:
            self.k1 = 0.85 - 0.05 * (fc * 10 - 280) / 70
            if self.k1 < 0.65: self.k1 = 0.65
        else:
            self.k1 = 0.85
        self.k2 = self.k1 / 2.0
        self.k3 = 0.85

        # Check print out
        """
        print("# Column obj")
        print("##  fc, fy, n  =", self.fc,", ",self.fy,", ",self.yn)
        print("##  B x D   =", self.b," x ",self.dd)
        print("## ",len(self.xs),"-D",self.dia)
        """

    ########################################################################
    # for deep learning
    def feature(self):

        return\
            self.six, self.siy,\
            self.szx, self.szy, self.ag

    ########################################################################
    # for set_column
    def bar_pos(self,nx,ny,dtx,dty):

        delx = ( self.b   - 2.0 * dtx ) / (nx-1.0)
        dely = ( self.dd  - 2.0 * dty ) / (ny-1.0)

        for i in range(0,nx):
            self.xs.append( dtx + i * delx )
            self.ys.append( dty )

        for i in range(0,nx):
            self.xs.append( dtx + i * delx )
            self.ys.append( self.dd - dty )

        for i in range(0,ny-2):
            self.xs.append( dtx )
            self.ys.append( dty + (i+1) * dely )

        for i in range(0,ny-2):
            self.xs.append( self.b - dtx )
            self.ys.append( dty + (i+1) * dely )


    ########################################################################
    # 中立軸の位置から軸力を求める。
    def narc(self, b, dd, xn, fc, ft, pos, area):

        # init
        ####################
        d = np.max(pos)
        dc = np.min(pos)
        aw = area

        csigc = fc
        ssigc = self.yn * fc * (xn - dc) / xn
        ssigt = self.yn * fc * (xn - d) / xn

        if abs(ssigt) > abs(ssigc):
            if abs(ssigt) > ft:
              if ssigt < 0  : ssigt = -ft
              if ssigt >= 0 : ssigt = ft
              csigc = ssigt * xn
              csigc = csigc / self.yn / (xn - d)
              ssigc = self.yn * csigc * (xn - dc) / xn
        else:
            if abs(ssigc) > ft :
              if ssigc < 0 : ssigc = -ft
              if ssigc >= 0 : ssigc = ft
              csigc = ssigc * xn
              csigc = csigc / self.yn / (xn - dc)
              ssigt = self.yn * csigc * (xn - d) / xn

        # Concrete
        nac = 1 / 2 * b * csigc * xn
        if xn > dd:
            csigt = (xn-dd)/xn * csigc
            #nac = 0.5 * b * csigc * dd
            nac = 0.5 * ( csigc + csigt ) * b * dd


        # Steel Bar
        nas = 0.0
        for i in range(0,len(pos)):
            nas = nas + aw * self.yn * csigc * ( xn - pos[i] ) / xn

        # result
        return (nac + nas)/10**3


    ########################################################################
    # 中立軸から曲げモーメントを求める。
    def marc(self, b, dd, xn, fc, ft, pos, area, nn):

        # init
        ####################
        d = np.max(pos)
        dc = np.min(pos)
        aw = area

        csigc = fc
        ssigc = self.yn * fc * (xn - dc) / xn
        ssigt = self.yn * fc * (xn - d) / xn

        if abs(ssigt) > abs(ssigc) :
            if abs(ssigt) > ft :
                if ssigt < 0 : ssigt = -ft
                if ssigt >= 0 : ssigt = ft
                csigc = ssigt * xn
                csigc = csigc / self.yn / (xn - d)
                ssigc = self.yn * csigc * (xn - dc) / xn
        else:
            if abs(ssigc) > ft :
                if ssigc < 0 : ssigc = -ft
                if ssigc >= 0 : ssigc = ft
                csigc = ssigc * xn
                csigc = csigc / self.yn / (xn - dc)
                ssigt = self.yn * csigc * (xn - d) / xn

        # Concete
        cmc = 1.0 / 6.0 * b * csigc * xn ** 2
        if xn > dd:
            csigt = (xn-dd)/xn * csigc
            #print( csigc, csigt )
            cmc = 1.0 / 6.0 * b * ( csigc - csigt ) * dd ** 2 + 1.0/2.0 * csigt * b * dd**2

        smt = 0.0
        for i in range(0,len(pos)):
            smt = smt + ( aw * self.yn * csigc * (xn - pos[i]) / xn) * pos[i]

        #print(smw,smt,smc,nn)
        #print(nn,-(cmc + smt + 2.0 * smw + smc - nn * self.dd / 2.0 ) / 10**6)
        aaa = -(cmc + smt - nn * dd / 2.0 )
        return aaa/10**6


    ########################################################################
    # *** M - N
    # 軸力から許容曲げモーメントを求める。
    #       Allowable Bending Moment
    def mnarc(self, b, dd, fc, ft, pos, aw, nn):

        xn1 = 0.0000001
        xn2 = 1.0*self.dd

        kk1 = self.narc(b, dd, xn1, fc, ft, pos, aw) - nn
        kk2 = self.narc(b, dd, xn2, fc, ft, pos, aw) - nn

        for i in range(1,10000):

            kk1 = self.narc(b,dd,xn1, fc, ft, pos, aw) - nn
            kk2 = self.narc(b,dd,xn2, fc, ft, pos, aw) - nn

            if abs(kk2) < self.eps : break;
            xn = xn2
            xn2 = (kk2 * xn1 - kk1 * xn2) / (kk2 - kk1)
            xn1 = xn

        xn = xn2
        aa = self.marc(b,dd,xn, fc, ft, pos, aw, nn * 1000)
        if xn > dd :
            aa = 0.0

        #print("Cal:i=",i,"xn",xn)
        return aa

    ########################################################################
    # 許容曲げ
    def ma(self, direction, fc, ft, nn):

        if direction == "X":
            ma = self.mnarc(self.dd, self.b, fc, ft, self.xs, self.area, nn)
        elif direction == "Y":
            ma = self.mnarc(self.b, self.dd, fc, ft, self.ys, self.area, nn)

        else:
            print("err. ma")

        return ma

    ########################################################################
    # MN Curve
    def mnaGen(self,direction,fc,ft,div):

        #ndiv = 10
        # init
        ####################
        na = []
        ma = []

        #xnmax = 1.0 * min( self.b,self.dd )
        #xnmax = 2.0 * min( self.b,self.dd )
        xnmax = 2.0 * min( self.b,self.dd )
        xnmin = 0.001 *  min( self.b,self.dd )
        #xnmin = max(self.dtx,self.dty)+1.0
        delxn = (xnmax - xnmin ) / div


        # dir.
        ####################
        pos_a = []
        if direction == "X":
            pos_a = self.xs
            b = self.dd
            dd = self.b
        elif direction == "Y":
            pos_a = self.ys
            b = self.b
            dd = self.dd
        else:
            print("err. ma")


        # analysis
        ####################
        na.append( self.nca_t(ft)/10**3 )
        ma.append( 0.0)

        for i in range(0,div):

            xn = xnmin + i * delxn
            natmp = self.narc( b, dd, xn, fc, ft, pos_a, self.area)
            na.append( natmp )
            ma.append( self.marc( b, dd, xn, fc, ft, pos_a, self.area, natmp*1000.0) )


        na.append( self.nca_c(fc,ft)/10**3 )
        ma.append( 0.0)

        na = np.array(na)
        ma = np.array(ma)

        # plot
        ####################
        if __name__ == '__main__':
            plt.plot( ma,na, c='black', linewidth=0.5 ,marker=".",label='Allowable')
            plt.grid()
            plt.show()


        return na,ma

    ########################################################################
    # 短期軸耐力
    def nca_c(self,fc,ft):
    # fc: コンクリート圧縮許容応力度 N/mm2
    # ft: 鉄筋圧縮許容応力度 N/mm2
        #nca_c = ( self.b * self.dd - self.ag) *  fc + self.ag * ft
        sigc = fc
        sigt = self.yn * sigc
        if sigt > ft :
            sigc = sigt/self.yn
            sigt = ft
        nca_c = ( self.b * self.dd -self.ag ) *  sigc + self.ag * sigt
        return nca_c

    def nca_t(self,ft):
        nca_t = - self.ag * ft
        return nca_t

    ########################################################################
    ########################################################################
    ########################################################################
    ########################################################################
    # ここから終局です
    ########################################################################
    ########################################################################
    ########################################################################
    ########################################################################
    ########################################################################
    ########################################################################


    ########################################################################
    # 終局引張強度
    def nu_t(self):

        nu_t = -self.ag * self.fy
        return nu_t

    def nu_c(self):

        nu_c = self.k3 * self.fc * (self.b * self.dd - self.ag) + self.ag * self.fy
        return nu_c

    ########################################################################
    # 設計軸力に対する曲げ耐力
    # by newton method
    def mnuaci(self,direction,nd):

        # MN
        # RC断面の終局曲げ耐力の算定
        # ACI

        nn = nd * 10**3

        fc = self.fc
        ee = 2.05*10**5
        sigy = self.fy

        # init
        ####################
        # aci stress block set
        eu = self.eu
        k1 = self.k1
        k2 = self.k2
        k3 = self.k3

        # dir.
        ####################
        if direction == "X":

            b = self.dd
            dd = self.b
            pos = self.xs

        elif direction == "Y":

            b = self.b
            dd = self.dd
            pos = self.ys

        else:
            print("err. mnuGen")

        #'収斂計算

        EPS = 10 ** (-3)
        xn1 = 0.00000001
        xn2 = 2.0*dd

        kk1 = self.nuaci(xn1, fc, eu, k1, k2, k3, ee, sigy,\
                         b, dd, pos, self.area) - nn
        kk2 = self.nuaci(xn2, fc, eu, k1, k2, k3, ee, sigy,\
                         b, dd, pos, self.area) - nn


        for i in range( 0 , 10000):
            kk1 = self.nuaci(xn1, fc, eu, k1, k2, k3, ee, sigy,\
                             b, dd, pos, self.area) - nn
            kk2 = self.nuaci(xn2, fc, eu, k1, k2, k3, ee, sigy,\
                             b, dd, pos, self.area) - nn
            if abs(kk2) < EPS: break;
            xn = xn2
            xn2 = (kk2 * xn1 - kk1 * xn2) / (kk2 - kk1)
            if xn2 < 0.0: xn2 = xn1 + 2.0
            xn1 = xn


        xn = xn2
        mnuaci = -self.muaci(xn, fc, eu, k1, k2, k3, ee, sigy,\
                             b, dd, pos, self.area, nn) / 10 ** 6

        if k1 * xn > 2.0*dd: mnuaci = 0.0
        if xn < 0.0: mnuaci = 0.0

        return mnuaci

    ########################################################################
    # 設計軸力に対する曲げ耐力
    # ２分法で求める 2022.03.07
    def mnuaci_twoM(self,direction,nd):

        # MN
        # RC断面の終局曲げ耐力の算定
        # ACI

        nn = nd * 10**3

        fc = self.fc
        ee = 2.05*10**5
        sigy = self.fy

        # init
        ####################
        # aci stress block set
        eu = self.eu
        k1 = self.k1
        k2 = self.k2
        k3 = self.k3

        # dir.
        ####################
        if direction == "X":

            b = self.dd
            dd = self.b
            pos = self.xs

        elif direction == "Y":

            b = self.b
            dd = self.dd
            pos = self.ys

        else:
            print("err. mnuGen")

        #'収斂計算

        EPS = 10 ** (-3)
        xn1 = 0.00000001
        xn2 = 2.0*dd

        kk1 = self.nuaci(xn1, fc, eu, k1, k2, k3, ee, sigy,\
                         b, dd, pos, self.area) - nn
        kk2 = self.nuaci(xn2, fc, eu, k1, k2, k3, ee, sigy,\
                         b, dd, pos, self.area) - nn


        for i in range( 0 , 10000):

            xn = 0.5 * ( xn1 + xn2 )

            kk1 = self.nuaci(xn1, fc, eu, k1, k2, k3, ee, sigy,\
                             b, dd, pos, self.area) - nn
            kk2 = self.nuaci(xn2, fc, eu, k1, k2, k3, ee, sigy,\
                             b, dd, pos, self.area) - nn

            kk  = self.nuaci(xn, fc, eu, k1, k2, k3, ee, sigy,\
                             b, dd, pos, self.area) - nn

            if abs(kk) < EPS: break;
            if kk*kk1 > 0:
                xn1 = xn
            else:
                xn2 = xn

        xn0 = xn
        mnuaci = -self.muaci(xn0, fc, eu, k1, k2, k3, ee, sigy,\
                             b, dd, pos, self.area, nn)\
            / 10 ** 6

        if k1 * xn > 2.0*dd: mnuaci = 0.0
        if xn < 0.0: mnuaci = 0.0

        return mnuaci

    ########################################################################
    # output to exel data
    def mn_result_xlsx(self,div,path,sheet_name,wb):
    # div: divided num
    # path: output file path
    # sheet_name: sheet name in excel

        # solve
        ####################
        nx, mx = self.mnuGen("X",div)
        ny, my = self.mnuGen("Y",div)
        # Permissible Stress
        nxa, mxa = self.mnaGen("X",self.sfc,self.sfy,div)
        nya, mya = self.mnaGen("Y",self.sfc,self.sfy,div)
        # fc,fyの設定が外部からアクセスできない

        #wb = openpyxl.Workbook(path)
        #wb = openpyxl.load_workbook(path)
        #ws1 = wb.active
        #ws1.title = sheet_name
        ws1 = wb.create_sheet(sheet_name)

        header1 = ['spec', 'mx','nx','my','ny','mxa','nxa','mya','nya']
        for i in range(len(header1)):
            ws1.cell(row = 1, column = i + 1).value = header1[i]

        for i in range(len(nx)):
            # Ultimate
            ws1.cell(row= i+2, column = 2 ).value = mx[i]
            ws1.cell(row= i+2, column = 3 ).value = nx[i]
            ws1.cell(row= i+2, column = 4 ).value = my[i]
            ws1.cell(row= i+2, column = 5 ).value = ny[i]
            # Permissible
            ws1.cell(row= i+2, column = 6 ).value = mxa[i]
            ws1.cell(row= i+2, column = 7 ).value = nxa[i]
            ws1.cell(row= i+2, column = 8 ).value = mya[i]
            ws1.cell(row= i+2, column = 9 ).value = nya[i]

        # output specification
        ws1.cell(row= 2, column = 1 ).value = self.b
        ws1.cell(row= 3, column = 1 ).value = self.dd
        ws1.cell(row= 4, column = 1 ).value = self.fc
        ws1.cell(row= 5, column = 1 ).value = self.fy
        ws1.cell(row= 6, column = 1 ).value = self.dia
        ws1.cell(row= 7, column = 1 ).value = len(self.xs)

        wb.save(path)

    ########################################################################
    # make model figure by object
    def make_model_fig(self,name):

        fig2 = plt.figure(figsize=(4/2.54,4/2.54),tight_layout=True)
        ax3 = fig2.add_subplot(1,1,1)

        ax3.set_aspect('equal')
        ax3.axes.xaxis.set_ticks([])
        ax3.axes.yaxis.set_ticks([])
        ax3.axis("off")
        #ax3.set_aspect('equal', adjustable='box')
        # for concrete
        """
        fib = patches.Rectangle(xy=(0, 0), width=self.b, height=self.dd, \
                                linewidth="0.5", ec='black',\
                                color="None", alpha=0.5 )
        """
        fib = patches.Rectangle(xy=(0, 0), width=self.b, height=self.dd, \
                                linewidth="0.5", ec='black',\
                                color="None")
        ax3.add_patch(fib)
        # for steel bar
        c = []
        for i in range(0,len(self.xs)):
            c.append(\
                     patches.Circle(\
                                    xy=(self.xs[i],self.ys[i]),\
                                    radius=self.dia/2.0, ec='black',\
                                    fc='black' ))
        for i in range(0,len(self.xs)):
            ax3.add_patch(c[i])
        ax3.plot()

        fig2.savefig(\
                     "./db/"+name+"_model.png",dpi=300,transparent=True\
                     )
        plt.close(fig2)

    ########################################################################
    # make figure by excel
    def make_fig(self,input_file,name):

        df_mn = pd.read_excel(input_file, sheet_name=name, engine='openpyxl' )
        direction = "XY"

        fig = plt.figure(figsize=(12/2.54,6/2.54),tight_layout=True)

        if direction == "XY":
            limx = max(np.max(df_mn['mx']),np.max(df_mn['my']))*1.2
        elif direction == "X":
            limx = np.max(df_mn['mx'])*1.2
        elif direction == "Y":
            limx = np.max(df_mn['my'])*1.2
        else:
            print("Err. pls, input direction to aft_mn")

        if direction != "Y":
            ax = fig.add_subplot(1,2,1)

        if direction != "X":
            ax2 = fig.add_subplot(1,2,2)

        ########################################################################
        # for mun diagram
        if direction != "Y":
            #ax.set_xlabel('Bending Moment, kN.m',fontsize="8")
            #ax.set_ylabel('Axial Force, kN',fontsize="8")
            ax.set_xlim(0,limx)
            ax.grid()
            ax.axhline(y=0,color='black',linewidth=0.5,linestyle='--')
            ax.axvline(x=0,color='black',linewidth=0.5,linestyle='--')
            ax.spines['right'].set_visible(False)
            ax.spines['top'].set_visible(False)
            ax.tick_params(labelsize="8")
            #ax.set_title("X-dir.",fontsize="9")
            #plt.plot( mx,nx, 'black' ,marker=".",label='Ultimate')
            #ax.plot( df_mn['mx'],df_mn['nx'], 'black' ,label='_nolegend_')
            ax.plot( df_mn['mx'],df_mn['nx'], 'black' ,label='Capacity')
            ax.legend(fontsize=8)
            ax.legend('upper right')
            ax.get_legend().remove()

        ########################################################################
        # for mun diagram
        if direction != "X":
            #ax2.set_xlabel('Bending Moment, kN.m', loc='left',fontsize="8")
            #ax2.set_xlabel('Bending Moment, kN.m', fontsize="8")
            #ax2.set_ylabel('Axial Force, kN', fontsize="8",)
            #a.set_ylabel('xlabel loc='+str(loc[i])+'', loc=loc[i])
            ax2.set_xlim(0,limx)
            ax2.grid()
            ax2.axhline(y=0,color='black',linewidth=0.5,linestyle='--')
            ax2.axvline(x=0,color='black',linewidth=0.5,linestyle='--')
            ax2.spines['right'].set_visible(False)
            ax2.spines['top'].set_visible(False)
            ax2.tick_params(labelsize="8")
            #ax2.set_title("Y-dir.",fontsize="9")
            #ax2.plot( df_mn['my'],df_mn['ny'], 'black',label='_nolegend_')
            ax2.plot( df_mn['my'],df_mn['ny'], 'black',label='Capacity')
            ax2.legend(fontsize=8)
            ax2.legend('upper right')
            ax2.get_legend().remove()

        #plt.show()
        fig.savefig("./db/"+name+"_mn.png",dpi=300,transparent=True)
        #plt.clf()
        plt.close(fig)

    ########################################################################
    # plot by matplotlib
    def aft_mn(self, direction, div, path):
    # nx,mux: axial force & bending moment in x-dir.
    # ny,muy: axial force & bending moment in y-dir.
    # direction/ "X" : x-dir
    #            "Y" : y-dir
    #            "XY" : both dir

        if direction == "XY":
            nx, mx = self.mnuGen("X",div)
            ny, my = self.mnuGen("Y",div)
            limx = max(np.max(mx),np.max(my))*1.2
        elif direction == "X":
            nx, mx = self.mnuGen(direction,div)
            limx = np.max(mx)*1.2
        elif direction == "Y":
            ny, my = self.mnuGen(direction,div)
            limx = np.max(my)*1.2
        else:
            print("Err. pls, input direction to aft_mn")


        # plot
        ####################
        #fig = plt.figure(figsize=(8,3))
        #fig = plt.figure(figsize=(9,3))
        #fig = plt.figure(figsize=(18/2.54,6/2.54))
        fig = plt.figure(figsize=(12/2.54,6/2.54),tight_layout=True)
        fig2 = plt.figure(figsize=(4/2.54,4/2.54),tight_layout=True)
        #fig = plt.figure(figsize=(18/2.54,9/2.54))

        # define ax for fig
        """
        1   2  3
        4   5  6
        7   8  9
        10 11 12

        1 2 3 4
        5 6 7 8
        1   2  3  4  5  6  7  8  9 10
        11 12 11 12 13 14 15 16
        17 18 18 19 20 21 22 23

        1 2 3
        4 5 6
        """
        #ax = plt.axes()
        if direction != "Y":
            #ax = fig.add_subplot(3,3,(2,8))
            #ax = fig.add_subplot(1,8,(3,5))
            #ax = fig.add_subplot(1,3,2)
            ax = fig.add_subplot(1,2,1)

        if direction != "X":
            #ax2 = fig.add_subplot(3,3,(3,9))
            #ax2 = fig.add_subplot(1,8,(6,8))
            #ax2 = fig.add_subplot(1,3,3)
            ax2 = fig.add_subplot(1,2,2)

        #ax3 = fig.add_subplot(4,3,(1,7))
        #ax3 = fig.add_subplot(2,3,1)
        ax3 = fig2.add_subplot(1,1,1)

        ########################################################################
        # for mun diagram
        if direction != "Y":
            #ax.set_xlabel('Bending Moment, kN.m',fontsize="8")
            #ax.set_ylabel('Axial Force, kN',fontsize="8")
            ax.set_xlim(0,limx)
            ax.grid()
            ax.axhline(y=0,color='black',linewidth=0.5,linestyle='--')
            ax.axvline(x=0,color='black',linewidth=0.5,linestyle='--')
            ax.spines['right'].set_visible(False)
            ax.spines['top'].set_visible(False)
            ax.tick_params(labelsize="8")
            #ax.set_title("X-dir.",fontsize="9")
            #plt.plot( mx,nx, 'black' ,marker=".",label='Ultimate')
            ax.plot( mx,nx, 'black' ,label='Ultimate')
            ax.legend(fontsize=8)
            # test
            nd = 0.0
            ax.scatter(self.mnuaci("X",nd),nd,marker="D")
            nd = 0.3 * np.max(nx)
            ax.scatter(self.mnuaci_twoM("X",nd),nd,marker="D")
            nd = 0.75 * np.min(nx)
            ax.scatter(self.mnuaci_twoM("X",nd),nd,marker="D")
            nd = 0.7 * np.max(nx)
            ax.scatter(self.mnuaci_twoM("X",nd),nd,marker="D")
            #ax.set_aspect('square')

        ########################################################################
        # for mun diagram
        if direction != "X":
            #ax2.set_xlabel('Bending Moment, kN.m', loc='left',fontsize="8")
            #ax2.set_xlabel('Bending Moment, kN.m', fontsize="8")
            #ax2.set_ylabel('Axial Force, kN', fontsize="8",)
            #a.set_ylabel('xlabel loc='+str(loc[i])+'', loc=loc[i])
            ax2.set_xlim(0,limx)
            ax2.grid()
            ax2.axhline(y=0,color='black',linewidth=0.5,linestyle='--')
            ax2.axvline(x=0,color='black',linewidth=0.5,linestyle='--')
            ax2.spines['right'].set_visible(False)
            ax2.spines['top'].set_visible(False)
            ax2.tick_params(labelsize="8")
            #ax2.set_title("Y-dir.",fontsize="9")
            ax2.plot( my,ny, 'black',label='Ultimate')
            ax2.legend(fontsize=8)

        ####################
        # for column shape output


        #ax3.spines['right'].set_visible(False)
        #ax3.spines['top'].set_visible(False)
        #ax3.spines['left'].set_visible(False)
        #ax3.spines['bottom'].set_visible(False)
        #ax3.axis('scaled')
        ax3.set_aspect('equal')
        ax3.axes.xaxis.set_ticks([])
        ax3.axes.yaxis.set_ticks([])
        ax3.axis("off")
        #ax3.set_aspect('equal', adjustable='box')
        # for concrete
        """
        fib = patches.Rectangle(xy=(0, 0), width=self.b, height=self.dd, \
                                linewidth="0.5", ec='#000000', color="None",\
                                alpha=0.5 )
        """
        fib = patches.Rectangle(xy=(0, 0), width=self.b, height=self.dd, \
                                linewidth="0.5", ec='black', color="None",\
                                alpha=0.5 )
        ax3.add_patch(fib)
        # Steel Bar
        c = []
        for i in range(0,len(self.xs)):
            c.append(\
                     patches.Circle(\
                                    xy=(self.xs[i],self.ys[i]),\
                                    radius=self.dia/2.0, ec='black', fc='black' ))
        for i in range(0,len(self.xs)):
            ax3.add_patch(c[i])
        ax3.plot()

        #plt.tight_layout()
        #plt.show()
        fig.savefig(path+"_mn.png",dpi=300,transparent=True)
        fig2.savefig(path+"_model.png",dpi=300,transparent=True)

        #plt.clf()
        plt.close(fig)
        plt.close(fig2)


    ########################################################################
    # 終局MNカーブの作成
    #def mnuGen(self, fc, ee, sigy, b, dd, d, dc, at, ac, nn, nw, aw, code, dt2):
    #def mnaGen(self,direction,fc,ft,div):
    def mnuGen(self, direction,div):

        fc = self.fc
        ee = 2.05*10**5
        sigy = self.fy

        # init
        ####################
        # aci stress block set
        eu = self.eu
        k1 = self.k1
        k2 = self.k2
        k3 = self.k3

        # parameter set
        nu = []
        mu = []

        #xnmax = 2.0 * min( self.b,self.dd )
        xnmax = 2.0 * min( self.b,self.dd )
        xnmin = 0.05 *  min( self.b,self.dd )
        delxn = (xnmax - xnmin ) / div

        # dir.
        ####################
        if direction == "X":

            b = self.dd
            dd = self.b
            pos = self.xs

        elif direction == "Y":

            b = self.b
            dd = self.dd
            pos = self.ys

        else:
            print("err. mnuGen")


        # analysis
        ####################

        nu.append( self.nu_t())
        mu.append( 0.0 )

        for i in range(0,div):

            xn = xnmin + i * delxn

            tmpnu = self.nuaci(xn, fc, eu, k1, k2, k3,\
                               ee, sigy, b, dd, pos, self.area)
            nu.append(tmpnu)
            mu.append(\
                      -self.muaci(xn, fc, eu, k1, k2, k3,\
                                  ee, sigy, b, dd, pos, self.area, tmpnu)\
                      )
            #self.marc( xn, fc, ft, d, dc, self.atx, self.acx, self.nwx, self.aw, na[i]*1000.0) )

        nu.append( self.nu_c())
        mu.append( 0.0 )

        nu = np.array(nu)/10**3 # N to kN
        mu = np.array(mu)/10**6 # N.mm to kN.m

        return nu,mu


    ########################################################################
    # 鉄筋の復元力特性
    def sig2(self, ee, sigy, e):
        # Bi-Linearの鉄筋スケルトン
        #ひずみから応力度を算出します

        ey = sigy / ee

        if abs(e) < ey :
            sig2 = ee * e

        else:
            if e > 0.0 :
                sig2 = sigy
            else:
                sig2 = -sigy

        return sig2
    ########################################################################
    # ultimate strength
    # 中立軸位置から軸力を算出します
    # use sig2
    def nuaci(self, xn, fc, eu, k1, k2, k3, ee, sigy, b, dd, pos, area):
        #d, dc, at, ac, nw, aw, code, dt2):
        #d,dc,at,ac,nw,aw,code,dt2 -----> pos, area

        #sep :引張り鉄筋の歪
        sep = 0.01 # 使ってない、いつかね。
        c = k1 * xn

        # Concrete
        if c > dd :
            nuaci = k3 * b * dd * fc;
        else:
            nuaci = k1 * k3 * b * xn * fc

        # steel bar
        ns = 0.0
        for i in range(0,len(pos)):
            e = -eu/xn * pos[i] + eu
            if c > pos[i]:
                ns = ns + ( self.sig2(ee, sigy, e) - k3*fc ) * area
            else:
                ns = ns + self.sig2(ee, sigy, e)* area

        # axial force
        nuaci = nuaci + ns

        return nuaci

    ########################################################################
    # 中立軸位置から曲げモーメントを算出します。
    def muaci(self,xn, fc, eu, k1, k2, k3, ee, sigy, b, dd, pos, area, nn):

        #d, dc, at, ac, nw, aw, nn, code, dt2):

        #sep :引張り鉄筋の歪, 使用していない。いつかね。
        sep = 0.01
        c = k1*xn;

        # concrete
        if c > dd :
            muaci = 0.5* dd**2 * k3 * b * fc;
        else:
            muaci = k1 * k2 * k3 * b * fc * xn ** 2

        # steel bar

        ms = 0.0
        for i in range(0,len(pos)):
            e = -eu/ xn * pos[i] + eu
            if c > pos[i]:
                ms = ms + (self.sig2(ee, sigy, e) - k3*fc ) * area * pos[i]
            else:
                ms = ms + self.sig2(ee, sigy, e) * area * pos[i]

        # bending moment
        muaci = muaci + ms - nn * dd / 2

        return muaci

########################################################################
# End of Class

########################################################################
# coded by tsunoppy
# make figure to use excel output data by matplotlib
class Aft_mn():


    ########################################################################
    # init
    def __init__(self,input_file,name,direction,mdmax,ndmin,ndmax):


        # make data frame
        self.df_mn = pd.read_excel(input_file, sheet_name=name,\
                                   engine='openpyxl' )
        self.name = name

        # make limit design
        if mdmax == "*":
            self.limx = max(np.max(self.df_mn['mx']),\
                            np.max(self.df_mn['my']))*1.2
        else:
            self.limx = mdmax

        self.ndmin = ndmin
        self.ndmax = ndmax

    def spec(self):

        return self.df_mn['spec'][0:6]

    ########################################################################
    # make figure by excel & stress
    def make_fig(self,xnum,xtitle,nux,mux,ynum,ytitle,nuy,muy):

        direction = "XY"

        fig = plt.figure(figsize=(12/2.54,6/2.54),tight_layout=True)

        if direction != "Y":
            ax = fig.add_subplot(1,2,1)

        if direction != "X":
            ax2 = fig.add_subplot(1,2,2)

        ########################################################################
        # for mun diagram
        if direction != "Y":
            #ax.set_xlabel('Bending Moment, kN.m',fontsize="8")
            #ax.set_ylabel('Axial Force, kN',fontsize="8")
            ax.set_xlim(0,self.limx)
            if self.ndmin != "*" and self.ndmax != "*":
                ax.set_ylim(self.ndmin,self.ndmax)
            elif self.ndmin == "*" and self.ndmax != "*":
                ax.set_ylim(np.min(self.df_mn['nx']),self.ndmax)
            elif self.ndmin != "*" and self.ndmax == "*":
                ax.set_ylim(self.ndmin,np.max(self.df_mn['nx']))
            else:
                print("ylim:",self.ndmin,self.ndmax )

            ax.grid()
            ax.axhline(y=0,color='black',linewidth=0.5,linestyle='--')
            ax.axvline(x=0,color='black',linewidth=0.5,linestyle='--')
            ax.spines['right'].set_visible(False)
            ax.spines['top'].set_visible(False)
            ax.tick_params(labelsize="8")
            #ax.set_title("X-dir.",fontsize="9")
            #plt.plot( mx,nx, 'black' ,marker=".",label='Ultimate')
            """
            ax.plot( self.df_mn['mx'],self.df_mn['nx'],\
                     'black' ,label='_nolegend_')
            """
            # Ultimate
            ax.plot( self.df_mn['mx'],self.df_mn['nx'],\
                     'black' ,label='_nolegend')
            # Permissible
            ax.plot( self.df_mn['mxa'],self.df_mn['nxa'],\
                     'black' ,label='_nolegend')

            ################
            face = ["black","None"]
            if xnum != 0:
                for i in range(0,xnum):
                    ax.scatter( mux[i], nux[i], marker="o", s=8, \
                                facecolor=face[i], edgecolor='black', label=xtitle[i])

            ax.legend('upper right')
            ax.legend(fontsize=8)
            #ax.get_legend().remove()

        ########################################################################
        # for mun diagram
        if direction != "X":
            #ax2.set_xlabel('Bending Moment, kN.m', loc='left',fontsize="8")
            #ax2.set_xlabel('Bending Moment, kN.m', fontsize="8")
            #ax2.set_ylabel('Axial Force, kN', fontsize="8",)
            #a.set_ylabel('xlabel loc='+str(loc[i])+'', loc=loc[i])
            ax2.set_xlim(0,self.limx)
            if self.ndmin != "*" and self.ndmax != "*":
                ax2.set_ylim(self.ndmin,self.ndmax)
            elif self.ndmin == "*" and self.ndmax != "*":
                ax2.set_ylim(np.min(self.df_mn['nx']),self.ndmax)
            elif self.ndmin != "*" and self.ndmax == "*":
                ax2.set_ylim(self.ndmin,np.max(self.df_mn['nx']))
            else:
                print("ylim:",self.ndmin,self.ndmax )

            ax2.grid()
            ax2.axhline(y=0,color='black',linewidth=0.5,linestyle='--')
            ax2.axvline(x=0,color='black',linewidth=0.5,linestyle='--')
            ax2.spines['right'].set_visible(False)
            ax2.spines['top'].set_visible(False)
            ax2.tick_params(labelsize="8")
            #ax2.set_title("Y-dir.",fontsize="9")
            """
            ax2.plot( self.df_mn['my'],self.df_mn['ny'],\
                      'black',label='_nolegend_')
            """
            # Ultimate
            ax2.plot( self.df_mn['my'],self.df_mn['ny'],\
                      'black',label='_nolegend')
            # Permissible
            ax2.plot( self.df_mn['mya'],self.df_mn['nya'],\
                     'black' ,label='_nolegend')

            ################


            if ynum != 0:
                for i in range(0,ynum):
                    ax2.scatter( muy[i], nuy[i], marker="o", s=8, facecolor=face[i], edgecolor='black', label=ytitle[i])

            ax2.legend('upper right')
            ax2.legend(fontsize=8)
            #ax2.get_legend().remove()

        #plt.show()
        fig.savefig("./db/"+self.name+"_mn.png",dpi=300,transparent=True)
        #plt.clf()
        plt.close(fig)


########################################################################


if __name__ == '__main__':


    """
    fc = 60
    fy = 490.0
    b = 1200.0
    dd = 950.0
    nx = [9,3]
    ny = [7,3]
    dtx = [100.0,475.0]
    dty = [100.0,350.0]
    dia = 41
    """

    #
    fc = 30.
    fy = 390.
    b = 750.
    dd = 750.
    nx = [4]
    ny = [6]
    dtx = [100.0]
    dty = [100.0]
    dia = 29
    
    nn = 950*950*60/3/1000
    nn = 0

    #obj = col(fc,fy,b2,dd,nx,ny,dtx,dty,dia)
    #print( obj.mnarc(fc, ft, d, dc, at, ac, nn, nw, aw) )

    obj = col(fc,fy,b,dd,nx,ny,dtx,dty,dia)

    fcb = 2.0/3.0 * fc
    ftb = fy
    #print( obj.ma("Y",fcb,ftb,nn) )
    div = 30
    obj.mnaGen("X",fcb,ftb,div)

    #obj.mnuGen("X",div)
    #obj.mnuGen("Y",div)
    #obj.mnaGen("Y",fcb,ftb,div)
    #obj.mnuGen("Y",div)

    #print(obj.mnuaci("X",0.0))
    #obj.aft_mn("XY",div,'./test')

    """
    import report

    # cntl: control file
    cntl = "./"
    num = 3
    pathname = "./test_report.pdf"
    title = "mn"
    obj = report.Report(cntl)
    obj.create_pdf(num,pathname,title)
    """
